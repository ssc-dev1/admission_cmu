from odoo import fields, models
import io
import openpyxl
from openpyxl.utils import get_column_letter
from odoo.http import request, Controller, route, content_disposition
from odoo import fields


class ClassActivityExcel(Controller):
    @route(['/class/activity/download/excel/<int:faculty_id>/<int:term_id>'], methods=['Get'], type='http', auth="user")
    def class_activity_download(self, faculty_id, term_id, **kw):
        referrer = request.httprequest.referrer
        if referrer and referrer.startswith(request.httprequest.host_url):
            workbook = openpyxl.Workbook()
            worksheet = workbook.worksheets[0]
            workbook.active = worksheet
            domain = [('faculty_staff_id', '!=', False)]
            term_id = request.env['odoocms.academic.term'].sudo().search([('id', '=', term_id)]) if term_id != 0 else False
            faculty_id = request.env['odoocms.institute'].sudo().search([('id', '=', faculty_id)]) if faculty_id != 0 else False
            if faculty_id:
                domain.append(('primary_class_id.institute_id', '=', faculty_id.id))
            if term_id:
                domain.append(('primary_class_id.term_id', '=', term_id.id))

            faculty_class_ids = request.env['odoocms.class'].sudo().search(domain)
            data = []
            headers = set(faculty_class_ids.mapped(
                'assessment_component_ids').mapped('type_id'))
            for rec in faculty_class_ids:
                line = {
                    'Name': rec.faculty_staff_id.name,
                    'Program': rec.primary_class_id.batch_id.program_id.code,
                    'Course Code': rec.code.split('-')[0] if rec.code else '-',
                    'Course Title': rec.name,
                    'Sections': rec.primary_class_id.section_name or '-',
                }
                for header in headers:
                    assessments = rec.assessment_ids.filtered(
                        lambda x: x.assessment_component_id.type_id == header)
                    if assessments:
                        assessments_approved = assessments.filtered(
                            lambda x: x.is_approved)
                        line.update({
                            header.name: f"{ len(assessments - assessments_approved) } + {len(assessments_approved)} = {len(assessments)}"
                        })
                    else:
                        line.update({
                            header.name: f"{0}+{0}={0}"
                        })

                data.append(line)

            # Create a BytesIO object to store the Excel file in memory
            
            excel_file = io.BytesIO()
            start_col = 6
            first_iteration = True
            for header in headers:
                start_row = 1
                if first_iteration:
                    worksheet[f'A{start_row}'].value = 'Serial No'
                    worksheet[f'B{start_row}'].value = 'Name'
                    worksheet[f'C{start_row}'].value = 'Program'
                    worksheet[f'D{start_row}'].value = 'Course Code'
                    worksheet[f'E{start_row}'].value = 'Course Title'
                    worksheet[f'F{start_row}'].value = 'Sections'
                    
                column = get_column_letter(start_col)
                worksheet[f'{column}{1}'].value = header.name
                worksheet.column_dimensions[column].width = 13
                start_col += 1
                serial_no = 1
                for rec in data:
                    if first_iteration:
                        worksheet[f'A{start_row + 1}'].value = serial_no
                        worksheet.column_dimensions['A'].width = 13
                        worksheet[f'B{start_row + 1}'].value = rec.get('Name', '')
                        worksheet.column_dimensions['B'].width = 13
                        worksheet[f'C{start_row + 1}'].value = rec.get('Program', '')
                        worksheet.column_dimensions['C'].width = 13
                        worksheet[f'D{start_row + 1}'].value = rec.get('Course Code', '')
                        worksheet.column_dimensions['D'].width = 13
                        worksheet[f'E{start_row + 1}'].value = rec.get('Course Title', '')
                        worksheet.column_dimensions['E'].width = 13
                        worksheet[f'F{start_row + 1}'].value = rec.get('Sections', '')
                        worksheet.column_dimensions['F'].width = 13
                    
                    serial_no += 1
                    start_row += 1
                    worksheet[f'{column}{start_row}'].value = rec.get(header.name, '')
                first_iteration = False
                
            
            # Save the workbook to the BytesIO object
            workbook.save(excel_file)
            workbook.close()
            # Set the position of the BytesIO object to the beginning
            excel_file.seek(0)
            # Return the file for download
            file_name = f'activity_report-{fields.Datetime.today().date()}.xlsx'
            return request.make_response(
                excel_file.getvalue(),
                headers=[('Content-Type', 'application/octet-stream'),
                         ('Content-Disposition', content_disposition(file_name))]
            )


class ClassActivityReportWizard(models.TransientModel):
    _name = 'class.activity.report'
    _description = 'Class Activity Report'

    faculty_id = fields.Many2one(
        'odoocms.institute', string='Faculty/Institute')
    term_id = fields.Many2one('odoocms.academic.term', string='Term')

    def print_report(self):
        datas = {
            'faculty_id': self.faculty_id.id,
            'term_id': self.term_id.id,
        }
        return self.env.ref('odoocms_academic.action_class_activity_report').with_context(landscape=True).report_action(self, data=datas, config=False)

    def print_excel_report(self):
        # call controller & download excel report

        return {
            'type': 'ir.actions.act_url',
            'url': f'/class/activity/download/excel/{self.faculty_id.id or 0}/{self.term_id.id or 0}',
            'target': 'new', }
