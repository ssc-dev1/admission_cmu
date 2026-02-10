from odoo import http, SUPERUSER_ID
from odoo.http import request
# from odoo.addons.portal.controllers.portal import pager as portal_pager, CustomerPortal
from datetime import date
import json
from werkzeug.datastructures import FileStorage
import pdb
from odoo.exceptions import AccessError, UserError, RedirectWarning, ValidationError, Warning
import base64
import logging
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter, column_index_from_string
from odoo.modules.module import get_module_resource
import io
from copy import copy
import numpy as np
from openpyxl.styles import PatternFill, Border, Side
from collections import Counter


_logger = logging.getLogger(__name__)


class AdmissionMeritReport(http.Controller):

    # for excel report
    def merge_cell_value(self, worksheet, value, cell: str,):
        if cell in worksheet.merged_cells:
            worksheet.unmerge_cells(cell)
        else:
            worksheet.merge_cells(cell)
            worksheet.unmerge_cells(cell)
        worksheet[cell.split(':')[0]].value = value
        worksheet.merge_cells(cell)

    def apply_style(self, worksheet, col, row, col_style, row_style):
        col = get_column_letter(col)
        style_col = get_column_letter(col_style)
        worksheet[col +
                  str(row)].border = copy(worksheet[style_col + str(row_style)].border)
        worksheet[col +
                  str(row)].font = copy(worksheet[style_col + str(row_style)].font)
        worksheet[col + str(row)].alignment = copy(
            worksheet[style_col + str(row_style)].alignment)
        worksheet[col +
                  str(row)].fill = copy(worksheet[style_col + str(row_style)].fill)

    @http.route(['/final/merit/register/line/report/download/<int:merit_reg_id>'], type='http', methods=['POST', 'GET'], auth="user", csrf=False)
    def merit_regist_line_report(self, merit_reg_id=0, **kw):
        try:
            if request.httprequest.method == 'GET':
                merit_reg_id = request.env['odoocms.merit.registers'].sudo().search(
                    [('id', '=', merit_reg_id)])
                merit_reg_line_ids = merit_reg_id.merit_lines

            # summary_recs = 0

            # start_row = 24
            # start_column = 2
            # w_col, adj_col = 0, 0
            # w_lst, adj_fac_lst, grace_mark_lst, adj_lst, grade_lst = [], [], [], [], []

                prepare_header = tuple(list(set(merit_reg_line_ids.mapped('cbt_section_ids').mapped('name'))))
                file_content = get_module_resource(
                    'odoocms_merit_ucp', 'static/xls/merit_register_line.xlsx')
                excel_file = io.BytesIO(open(file_content, 'rb').read())
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.worksheets[0]
                workbook.active = worksheet
                header_ = 13

                for i in range(len(prepare_header)):
                    total_marks = merit_reg_line_ids.mapped('cbt_section_ids').filtered(lambda x: x.name == prepare_header[i])
                    if total_marks:
                        total_marks=total_marks[0].total_marks
                    worksheet[get_column_letter(header_) + str(1)].value = f"{prepare_header[i]} ({total_marks})" 
                    header_ += 1
                    worksheet[get_column_letter(header_) + str(1)].value = f"{prepare_header[i]} %" 
                    
                    header_ += 1

                row = 2
                col = 1
                worksheet[get_column_letter(1) + str(1)].value = "Merit No"
                worksheet[get_column_letter(2) + str(1)].value = "Ref No#"
                worksheet[get_column_letter(3) + str(1)].value = "Name"
                worksheet[get_column_letter(4) + str(1)].value = "Program"
                worksheet[get_column_letter(5) + str(1)].value = "Preference 2"
                worksheet[get_column_letter(6) + str(1)].value = "Preference 3"
                worksheet[get_column_letter(7) + str(1)].value = "Inter Status"
                worksheet[get_column_letter(8) + str(1)].value = "Matric %"
                worksheet[get_column_letter(9) + str(1)].value = "Inter %"
                worksheet[get_column_letter(10) + str(1)].value = "Pre Test %"
                worksheet[get_column_letter(11) + str(1)].value = "Pre Test Total"
                worksheet[get_column_letter(12) + str(1)].value = "Pre Test Obtained"
                # worksheet[get_column_letter(10) + str(1)].value = "Matric Total"
                # worksheet[get_column_letter(11) + str(1)].value = "Matric Obtained"
                # worksheet[get_column_letter(13) + str(1)].value = "Inter Total"
                # worksheet[get_column_letter(14) + str(1)].value = "Inter Obtained"
                # worksheet[get_column_letter(8) + str(1)].value = "AGGREGATE"
                # worksheet[get_column_letter(header_) + str(1)].value = "Pre Test %"
                # header_+= 1
                # worksheet[get_column_letter(header_) + str(1)].value = "Pre Test Total"
                # header_+= 1
                # worksheet[get_column_letter(header_) + str(1)].value = "Pre Test Obtained"
                worksheet[get_column_letter(header_) + str(1)].value = "CBT %"
                header_+= 1
                worksheet[get_column_letter(header_) + str(1)].value = "CBT Total"
                header_+= 1
                worksheet[get_column_letter(header_) + str(1)].value = "CBT Obtained"
                header_+= 1
                worksheet[get_column_letter(header_) + str(1)].value = "Religion"

                # worksheet[get_column_letter(19) + str(1)].value = "Pre Test Name"
                # worksheet[get_column_letter(20) + str(1)].value = "Pre Test %"
                # worksheet[get_column_letter(21) + str(1)].value = "Pre Test Marks"
                for student in merit_reg_line_ids:
                    worksheet[get_column_letter(col) + str(row)].value = student.merit_no
                    worksheet[get_column_letter(col+1) + str(row)].value = student.applicant_id.application_no
                    worksheet[get_column_letter(col+2) + str(row)].value = student.applicant_id.name
                    worksheet[get_column_letter(col+3) + str(row)].value = student.applicant_id.prefered_program_id.name
                    second_preference = student.applicant_id.preference_ids.filtered(lambda x:x.preference == 2)
                    third_preference = student.applicant_id.preference_ids.filtered(lambda x:x.preference == 3)
                    worksheet[get_column_letter(col+4) + str(row)].value = second_preference[0].program_id.name if second_preference else ''
                    worksheet[get_column_letter(col+5) + str(row)].value = third_preference[0].program_id.name if third_preference else ''
                    inter_status = student.applicant_id.applicant_academic_ids.filtered(lambda x:x.degree_name.year_age==12)
                    worksheet[get_column_letter(col+6) + str(row)].value = str(inter_status.result_status).upper() if inter_status else ''            

                    worksheet[get_column_letter(col+7) + str(row)].value = student.matric_marks_per
                    worksheet[get_column_letter(col+8) + str(row)].value = student.inter_marks_per
                    pre_test_percentage = (student.pretest_marks / student.applicant_id.pre_test_id.pre_test_total_marks) * 100 if student.applicant_id.pre_test_id.pre_test_total_marks else 0

                    worksheet[get_column_letter(col+9) + str(row)].value = pre_test_percentage or 0
                    worksheet[get_column_letter(col+10) + str(row)].value = student.applicant_id.pre_test_id.pre_test_total_marks or 0
                    worksheet[get_column_letter(col+11) + str(row)].value = student.pretest_marks or 0
                    # worksheet[get_column_letter(col+7) + str(row)].value = student.aggregate
                    # worksheet[get_column_letter(col+9) + str(row)].value = student.matric_total_marks
                    # worksheet[get_column_letter(col+10) + str(row)].value = student.matric_marks
                    # worksheet[get_column_letter(col+12) + str(row)].value = student.inter_total_marks
                    # worksheet[get_column_letter(col+13) + str(row)].value = student.inter_marks
                    # _logger.warning(student)
                    # worksheet[get_column_letter(col+18) + str(row)].value = student.pre_test_name
                    # worksheet[get_column_letter(col+19) + str(row)].value = student.pre_test_percentage
                    # worksheet[get_column_letter(col+20) + str(row)].value = student.pre_test_marks

                    header_col = 12
                    for i in range(len(prepare_header)):
                        section_cbt = student.cbt_section_ids.filtered(lambda x:x.name == prepare_header[i])
                        if section_cbt:
                            header_col += 1
                            section_cbt = section_cbt[0]
                            worksheet[get_column_letter(header_col) + str(row)].value = section_cbt.marks or 0
                            # header_col += 1
                            header_col += 1
                            worksheet[get_column_letter(header_col) + str(row)].value = ((section_cbt.marks or 0)/section_cbt.total_marks)*100
                        else:
                            worksheet[get_column_letter(header_col) + str(row)].value = ''
                            # header_col += 1
                            header_col += 1
                            worksheet[get_column_letter(header_col ) + str(row)].value = ''
                            

                    header_col += 1
                    worksheet[get_column_letter(header_col) + str(row)].value = student.cbt_percentage
                    header_col += 1
                    worksheet[get_column_letter(header_col) + str(row)].value = sum(student.mapped('cbt_section_ids').mapped('marks'))
                    header_col += 1
                    worksheet[get_column_letter(header_col) + str(row)].value = sum(student.mapped('cbt_section_ids').mapped('total_marks'))
                    header_col += 1
                    worksheet[get_column_letter(header_col) + str(row)].value = student.applicant_id.religion_id.name

                    
                        
                    row += 1
                # pattern_fill = PatternFill(fill_type='solid', start_color='008cb3')
                # pattern_fill2 = PatternFill(fill_type='solid', start_color='E6B0AA')
                # col_2_22_font = copy(worksheet[get_column_letter(2) + str(22)].font)
                # col_1_22_font = copy(worksheet[get_column_letter(1) + str(22)].font)

                # apply style - only formatting
                # row_length = len(students_data)
                # row_range = 22
                # for row in range(row_length + 2):
                #     assessment_cols = 6 + len(class_assessments)
                #     assessment_summarize_cols = summary_recs

                #     for col in range(1, assessment_cols + 1):
                #         if row_range >= 24:
                #             self.apply_style(worksheet, col + 1, str(row_range), 2, 24)
                #         else:
                #             if col >= 6:
                #                 self.apply_style(worksheet, col + 1, str(row_range), 7, 22)

                #     for col in range(1, assessment_summarize_cols + 1):
                #         if row_range in (22, 23,):
                #             worksheet[get_column_letter(col + assessment_cols) + str(row_range)].alignment = copy(worksheet[get_column_letter(1) + str(4)].alignment)
                #             worksheet[get_column_letter(col + assessment_cols) + str(row_range)].font = copy(worksheet[get_column_letter(1) + str(4)].font)
                #             worksheet[get_column_letter(col + assessment_cols) + str(row_range)].border = copy(worksheet[get_column_letter(7) + str(22)].border)
                #         else:
                #             self.apply_style(worksheet, col + assessment_cols, str(row_range), 2, 24)
                #         worksheet[get_column_letter(col + assessment_cols) + str(row_range)].fill = pattern_fill

                #     for col in range(1,6):
                #         if row_range in (22,23,):
                #             worksheet[get_column_letter(col + assessment_cols + assessment_summarize_cols) + str(row_range)].font = copy(worksheet[get_column_letter(1) + str(4)].font)
                #             worksheet[get_column_letter(col + assessment_cols + assessment_summarize_cols) + str(row_range)].alignment = copy(worksheet[get_column_letter(1) + str(4)].alignment)
                #             worksheet[get_column_letter(col + assessment_cols + assessment_summarize_cols) + str(row_range)].border = copy(worksheet[get_column_letter(7) + str(22)].border)
                #         else:
                #             self.apply_style(worksheet,col + assessment_cols + assessment_summarize_cols,str(row_range),2,24 )
                #         worksheet[get_column_letter(col + assessment_cols + assessment_summarize_cols) + str(row_range)].fill = pattern_fill2

                #     row_range += 1

                # column_assessment = start_column + 5
                # for line in class_assessments:
                #     worksheet.cell(start_row - 2, column_assessment).value = line.code
                #     worksheet.cell(start_row - 1, column_assessment, line.max_marks)
                #     column_assessment += 1

                # for line in assessment_components:
                #     worksheet.cell(start_row - 2, column_assessment, line.type_id.name)
                #     column_assessment += 1

                # start_row += 1
                # students_data = tuple(students_data)
                # for student in students_data:
                #     worksheet.cell(start_row,start_column,student.get('student', '').id_number)
                #     self.apply_style(worksheet,3,start_row,1,24)
                #     worksheet[get_column_letter(3) + str(start_row)].fill=copy(worksheet[get_column_letter(2) + str(24)].fill)
                #     self.merge_cell_value(worksheet=worksheet, value=student.get('student', '').name, cell=f"C{start_row}:F{start_row}")

                #     column_assessment = start_column + 5
                #     for k, assessment_value in student.get('assessment').items():
                #         worksheet.cell(start_row,column_assessment,assessment_value and assessment_value.obtained_marks or '-')
                #         column_assessment += 1

                #     for k, assessment_value in student.get('assessment_summarize').items():
                #         worksheet.cell(start_row,column_assessment,
                #             assessment_value and (round(assessment_value.percentage*(assessment_value.assessment_component_id.weightage or 1)/100, 2))) or '-'
                #         column_assessment += 1

                #     extra_assessment = ('Weightage','Adj Factor','Grace Marks','Adj','Grade',)
                #     for e in extra_assessment:
                #         if e == 'Weightage' and student.get('Grade','') != 'W':
                #             w_col = column_assessment
                #             w_lst.append(student.get(e,0))
                #         elif e == 'Adj Factor':
                #             adj_fac_lst.append(student.get(e,0))
                #         elif e == 'Grace Marks':
                #             grace_mark_lst.append(student.get(e,0))
                #         elif e == 'Adj' and student.get('Grade','') != 'W':
                #             adj_col = column_assessment
                #             adj_lst.append(student.get(e,0))
                #         elif e == 'Grade':
                #             grade_lst.append(student.get(e,''))

                #         worksheet.cell( 22,column_assessment,e)
                #         worksheet.cell( start_row,column_assessment,student.get(e,0))
                #         column_assessment += 1

                #     start_row += 1

                # student_assessments_summarize = dict(sorted(student_assessments_summarize.items()))
                # summarize_row = start_row
                # sumarize_col = 7
                # student_assessment_keys = student_data.get('assessment').keys()
                # for k in student_assessment_keys:
                #     v = student_assessments_summarize.get(k)
                #     worksheet.cell( summarize_row,sumarize_col,v.get('min'))
                #     worksheet.cell( summarize_row+1,sumarize_col,v.get('max'))
                #     worksheet.cell( summarize_row+2,sumarize_col,v.get('avg'))
                #     worksheet.cell( summarize_row+3,sumarize_col,v.get('std'))

                #     sumarize_col += 1

                # fil_row = summarize_row
                # assessments_summarize_list = ('min','max','avg','std',)
                # for fil in assessments_summarize_list:
                #     for col_fill in range(column_assessment - 2):
                #         self.apply_style(worksheet,col_fill+2,str(fil_row),1,22)
                #         worksheet[get_column_letter(col_fill + 2) + str(fil_row)].fill = pattern_fill
                #     fil_row += 1

                # for i in assessments_summarize_list:
                #     worksheet[get_column_letter(3) + str(summarize_row)].alignment = copy(worksheet[get_column_letter(1) + str(23)].alignment)
                #     self.merge_cell_value(worksheet=worksheet, value= i.upper(), cell=f"C{summarize_row}:F{summarize_row}")
                #     if i =='min':
                #         worksheet.cell(summarize_row,w_col,np.min(w_lst))
                #         worksheet.cell(summarize_row,adj_col,np.min(adj_lst))
                #         worksheet[get_column_letter(w_col) + str(summarize_row)].font = col_1_22_font
                #         worksheet[get_column_letter(adj_col) + str(summarize_row)].font = col_1_22_font
                #     elif i =='max':
                #         worksheet.cell(summarize_row,w_col,np.max(w_lst))
                #         worksheet.cell(summarize_row,adj_col,np.max(adj_lst))
                #         worksheet[get_column_letter(w_col) + str(summarize_row)].font = col_1_22_font
                #         worksheet[get_column_letter(adj_col) + str(summarize_row)].font = col_1_22_font
                #     elif i =='avg':
                #         worksheet.cell(summarize_row,w_col,np.average(w_lst))
                #         worksheet.cell(summarize_row,adj_col,np.average(adj_lst))
                #         worksheet[get_column_letter(w_col) + str(summarize_row)].font = col_1_22_font
                #         worksheet[get_column_letter(adj_col) + str(summarize_row)].font = col_1_22_font
                #     elif i =='std':
                #         worksheet.cell(summarize_row,w_col,np.std(w_lst))
                #         worksheet.cell(summarize_row,adj_col,np.std(adj_lst))
                #         worksheet[get_column_letter(w_col) + str(summarize_row)].font = col_1_22_font
                #         worksheet[get_column_letter(adj_col) + str(summarize_row)].font = col_1_22_font
                #     summarize_row += 1

                # section_name = class_id.primary_class_id.section_id and class_id.primary_class_id.section_id.name or ''
                # self.merge_cell_value(worksheet,class_id.primary_class_id.institute_id.name,f'M{3}:AC{4}')
                # self.merge_cell_value(worksheet,grade_class_id.code,f'P{6}:W{6}')
                # self.merge_cell_value(worksheet,grade_class_id.name,f'P{7}:W{7}')
                # self.merge_cell_value(worksheet,section_name,f'P{8}:W{8}')
                # self.merge_cell_value(worksheet,grade_class_id.grade_staff_id.name,f'P{9}:W{9}')

                # # # average all table
                # row = 4
                # assessment_components = class_id.assessment_component_ids
                # for assessment_component in assessment_components:
                #     self.apply_style(worksheet,2,row,2,24 )
                #     self.merge_cell_value(worksheet,assessment_component.type_id.name,f'B{row}:C{row}')
                #     self.apply_style(worksheet,4,row,2,24 )
                #     if assessment_component.consideration_avg:
                #         worksheet[get_column_letter(4) + str(row)].value = 'Avg All'
                #     elif assessment_component.consideration_top:
                #         worksheet[get_column_letter(4) + str(row)].value = 'Top ' + str(assessment_component.best)
                #     self.apply_style(worksheet,5,row,2,24)
                #     worksheet[get_column_letter(5) + str(row)].value = assessment_component.weightage or 0
                #     row += 1

                # grade_count = dict(Counter(grade_lst))
                # grade = ['A','A-','B+','B','B-','C+','C','C-','D+','D','W','F','I','P','S','US']
                # row=4
                # for k in grade:
                #     if int(grade_count.get(k,'0')) > 0:
                #         self.apply_style(worksheet,7,row,2,22)
                #         self.apply_style(worksheet,9,row,2,22)
                #         self.merge_cell_value(worksheet,k.upper(),f'G{row}:H{row}')
                #         self.merge_cell_value(worksheet,grade_count.get(k,'0'),f'I{row}:J{row}')
                #         row += 1

                buffer = io.BytesIO()
                workbook.save(buffer)
                headers = [
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', 'attachment; filename=merit_register_line.xlsx'),
                    ('Content-Length', len(buffer.getvalue()))
                ]
                return request.make_response(buffer.getvalue(), headers=headers)

        except Exception as e:
            print(e)
            _logger.exception(f'{e}')

            values = {
                'error_message': f"{e}"
            }
            return json.dumps(values)
            # return http.request.render('odoocms_web.portal_error', values)
